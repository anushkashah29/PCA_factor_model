"""
Treasury Yield PCA Factor Model & VaR Calculator
=================================================
Steps:
  1. Fetch CMT yields from FRED (3M,6M,12M,2Y,3Y,5Y,7Y,10Y,20Y,30Y)
  2. Clean data: drop weekends, US federal holidays, and NaN rows
  3. Compute daily yield changes; export raw yields + changes to Excel
  4. Covariance matrix → eigendecomposition
  5. Extract 3 PCs (level, slope, curvature)
  6. DV01 for 2Y & 10Y treasury bond portfolio → factor risk weights
  7. Historical simulation VaR: 99% confidence, 10-day horizon
"""

import os
import sys
import requests
import numpy as np
import pandas as pd
from scipy import stats
from datetime import date
import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt
import matplotlib.ticker as mticker
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import (
    PatternFill, Font, Alignment, Border, Side, numbers
)
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.chart.series import SeriesLabel

# ──────────────────────────────────────────────────────────────────────────────
# CONFIGURATION
# ──────────────────────────────────────────────────────────────────────────────
FRED_API_KEY = "0a7bcf57e2f38f50b9f230c76a4ff850"
FRED_BASE    = "https://api.stlouisfed.org/fred/series/observations"
START_DATE   = "2007-01-01"
END_DATE     = date.today().isoformat()
OUTPUT_FILE  = "treasury_pca_var.xlsx"

# FRED series IDs for CMT yields
SERIES = {
    "3M":  "DGS3MO",
    "6M":  "DGS6MO",
    "1Y":  "DGS1",
    "2Y":  "DGS2",
    "3Y":  "DGS3",
    "5Y":  "DGS5",
    "7Y":  "DGS7",
    "10Y": "DGS10",
    "20Y": "DGS20",
    "30Y": "DGS30",
}
TENORS   = list(SERIES.keys())   # ordered maturities
TENOR_YR = [0.25, 0.5, 1, 2, 3, 5, 7, 10, 20, 30]  # numeric years

# Portfolio: face values (USD)
PORTFOLIO = {
    "2Y":  10_000_000,   # $10 M notional 2Y T-note
    "10Y": 10_000_000,   # $10 M notional 10Y T-note
}
COUPON_RATES = {"2Y": 0.045, "10Y": 0.042}   # approximate current coupons

# VaR parameters
CONFIDENCE  = 0.99
HORIZON     = 10   # days
N_SIM       = 10_000

# US Federal Holidays (approximation via pandas USFederalHolidayCalendar)
from pandas.tseries.holiday import USFederalHolidayCalendar

def us_holidays(start, end):
    cal = USFederalHolidayCalendar()
    return cal.holidays(start=start, end=end)


# ──────────────────────────────────────────────────────────────────────────────
# STEP 1 – FETCH DATA
# ──────────────────────────────────────────────────────────────────────────────
def fetch_series(series_id: str) -> pd.Series:
    params = {
        "series_id":       series_id,
        "api_key":         FRED_API_KEY,
        "file_type":       "json",
        "observation_start": START_DATE,
        "observation_end":   END_DATE,
    }
    r = requests.get(FRED_BASE, params=params, timeout=30)
    r.raise_for_status()
    obs = r.json()["observations"]
    s = pd.Series(
        {o["date"]: float(o["value"]) if o["value"] != "." else np.nan
         for o in obs},
        name=series_id,
        dtype=float,
    )
    s.index = pd.to_datetime(s.index)
    return s


def fetch_all() -> pd.DataFrame:
    frames = {}
    for label, sid in SERIES.items():
        print(f"  Fetching {label:4s} ({sid}) …", end="", flush=True)
        s = fetch_series(sid)
        frames[label] = s
        print(f" {len(s)} obs")
    df = pd.DataFrame(frames)
    df.index.name = "Date"
    return df


# ──────────────────────────────────────────────────────────────────────────────
# STEP 2 – CLEAN DATA
# ──────────────────────────────────────────────────────────────────────────────
def clean(df: pd.DataFrame) -> pd.DataFrame:
    # Drop weekends (FRED should already exclude, but be safe)
    df = df[df.index.dayofweek < 5].copy()
    # Drop US federal holidays
    holidays = us_holidays(df.index.min(), df.index.max())
    df = df[~df.index.isin(holidays)]
    # Drop rows where ALL tenors are NaN (sparse dates)
    df = df.dropna(how="all")
    # Forward-fill isolated NaN cells (stale quotes on short-holidays for single tenors)
    df = df.ffill().dropna()
    return df.sort_index()


# ──────────────────────────────────────────────────────────────────────────────
# STEP 3 – DAILY YIELD CHANGES (in basis points)
# ──────────────────────────────────────────────────────────────────────────────
def yield_changes(yields: pd.DataFrame) -> pd.DataFrame:
    changes = yields.diff().dropna() * 100   # convert % → bps
    return changes


# ──────────────────────────────────────────────────────────────────────────────
# STEP 4 – COVARIANCE & EIGENDECOMPOSITION
# ──────────────────────────────────────────────────────────────────────────────
def eigen(changes: pd.DataFrame):
    """Return (eigenvalues, eigenvectors, cov_matrix) sorted desc by eigenvalue."""
    Sigma = np.cov(changes.values.T)         # (n_tenors, n_tenors)
    vals, vecs = np.linalg.eigh(Sigma)       # eigh for symmetric → real, sorted asc
    # Sort descending
    idx  = np.argsort(vals)[::-1]
    vals = vals[idx]
    vecs = vecs[:, idx]                      # columns are eigenvectors
    return vals, vecs, Sigma


# ──────────────────────────────────────────────────────────────────────────────
# STEP 5 – PRINCIPAL COMPONENTS
# ──────────────────────────────────────────────────────────────────────────────
def compute_pcs(changes: pd.DataFrame, vecs: np.ndarray) -> pd.DataFrame:
    X  = changes.values                          # (T, n_tenors)
    pc = X @ vecs[:, :3]                         # (T, 3)
    return pd.DataFrame(pc, index=changes.index,
                        columns=["PC1_Level", "PC2_Slope", "PC3_Curvature"])


# ──────────────────────────────────────────────────────────────────────────────
# STEP 6 – DV01 & FACTOR RISK WEIGHTS
# ──────────────────────────────────────────────────────────────────────────────
def bond_dv01(face: float, coupon: float, ytm: float, maturity_yrs: float,
              freq: int = 2) -> float:
    """Dollar value of 1 bp shift (in USD) for a fixed-coupon bond."""
    n   = int(maturity_yrs * freq)
    c   = coupon / freq * face
    r   = ytm / freq
    pv_func = lambda y: sum(c / (1 + y)**t for t in range(1, n + 1)) + face / (1 + y)**n
    price  = pv_func(r)
    price_up = pv_func(r + 0.0001 / freq)   # +1 bp / freq
    dv01   = (price - price_up)              # negative → bond price falls when yield rises
    return dv01


def compute_dv01s(yields_last: pd.Series) -> dict:
    """DV01 for each portfolio bond using last observed YTM."""
    dv01s = {}
    for tenor, face in PORTFOLIO.items():
        ytm     = yields_last[tenor] / 100      # % → decimal
        coupon  = COUPON_RATES[tenor]
        mat     = TENOR_YR[TENORS.index(tenor)]
        dv01s[tenor] = bond_dv01(face, coupon, ytm, mat)
    return dv01s


def factor_risk_weights(dv01s: dict, vecs: np.ndarray) -> np.ndarray:
    """
    Portfolio DV01 vector across all tenors (USD per bp at each tenor),
    projected onto the first 3 eigenvectors.
    Returns array shape (3,): dollar sensitivity per unit of each PC.
    """
    # Sparse DV01 vector across all tenors
    dv01_vec = np.zeros(len(TENORS))
    for tenor, dv01 in dv01s.items():
        idx = TENORS.index(tenor)
        dv01_vec[idx] = dv01
    # Risk weight for each PC: w_k = dv01_vec · e_k
    weights = vecs[:, :3].T @ dv01_vec    # shape (3,)
    return weights, dv01_vec


# ──────────────────────────────────────────────────────────────────────────────
# STEP 7 – VaR: Three Methods Compared
# ──────────────────────────────────────────────────────────────────────────────
def compute_var(pcs: pd.DataFrame, weights: np.ndarray,
                confidence: float = 0.99, horizon: int = 10,
                n_sim: int = 10_000) -> dict:
    """
    Three VaR methods — the first two assume normality and will agree closely;
    the third uses the actual empirical distribution and will differ whenever
    the actual PC changes have fat tails (excess kurtosis > 0).

    Method 1 — Analytical (closed-form):
        VaR = z * sqrt( sum( (w_k * sigma_k * sqrt(T))^2 ) )
        Normal assumption, exact formula.

    Method 2 — Parametric Monte Carlo (normal draws):
        Draw PC moves from N(0, sigma_k * sqrt(T)), compute P&L, take percentile.
        Also normal assumption. Will match Method 1 to within sampling error.
        Exists here only to demonstrate the convergence to Method 1.

    Method 3 — True Historical Simulation:
        Use actual observed daily PC scores from the full history.
        Scale each day's P&L by sqrt(T) (square-root-of-time rule).
        No distribution assumption — uses the real empirical tail directly.
        Captures GFC, COVID, and Fed-hiking spikes that normality misses.
    """
    pc_std = pcs.std().values                    # (3,) 1-day std of each PC (bps)
    sqrt_T = np.sqrt(horizon)

    # ── Method 1: Analytical ──────────────────────────────────────────────────
    var_analytical = stats.norm.ppf(confidence) * np.sqrt(
        sum((weights[k] * pc_std[k] * sqrt_T) ** 2 for k in range(3))
    )

    # ── Method 2: Parametric Monte Carlo (normal draws) ───────────────────────
    np.random.seed(42)
    sim_moves = np.zeros((n_sim, 3))
    for k in range(3):
        sim_moves[:, k] = np.random.normal(0, pc_std[k] * sqrt_T, n_sim)
    pnl_mc  = sim_moves @ weights
    var_mc  = abs(np.percentile(pnl_mc, (1 - confidence) * 100))

    # ── Method 3: True Historical Simulation ─────────────────────────────────
    # Each row of pcs is one day's actual observed PC moves.
    # P&L_daily = w1*PC1 + w2*PC2 + w3*PC3  (USD per day, using real data)
    # Scale to horizon via sqrt(T) — same assumption as Methods 1 & 2, but
    # the distribution of the input data is now empirical, not forced-normal.
    # Negate: eigh() orients PC1 loadings all-negative; yield fall => positive PC1
    # but negative pcs@weights => inverted P&L sign for a long bond portfolio.
    # Fix: long bonds gain when yields fall, so negate to get correct sign.
    daily_pnl       = -(pcs.values @ weights)        # (T,) sign-corrected daily P&L
    pnl_hist_10d    = daily_pnl * sqrt_T             # scaled to 10-day horizon
    var_hist        = abs(np.percentile(pnl_hist_10d, (1 - confidence) * 100))

    # Empirical distribution stats of actual daily P&L (explains gap vs analytical)
    pnl_skew = float(stats.skew(daily_pnl))
    pnl_kurt = float(stats.kurtosis(daily_pnl))     # excess kurtosis (0 = normal)

    # PC-level standalone contributions (historical simulation)
    contrib_hist = []
    for k in range(3):
        pnl_k = -(pcs.values[:, k] * weights[k]) * sqrt_T
        contrib_hist.append(abs(np.percentile(pnl_k, (1 - confidence) * 100)))

    return {
        "var_analytical_usd": var_analytical,
        "var_mc_usd":         var_mc,
        "var_hist_usd":       var_hist,
        "pnl_hist_10d":       pnl_hist_10d,
        "pnl_hist_dates":     pcs.index,
        "daily_pnl":          daily_pnl,
        "pnl_mc":             pnl_mc,
        "pc_std_1d":          pc_std,
        "pc_std_10d":         pc_std * sqrt_T,
        "contrib_hist":       contrib_hist,
        "pnl_skew":           pnl_skew,
        "pnl_kurt":           pnl_kurt,
        "sim_moves":          sim_moves,
        "weights":            weights,
    }


# ──────────────────────────────────────────────────────────────────────────────
# EXCEL OUTPUT
# ──────────────────────────────────────────────────────────────────────────────
HDR_FILL  = PatternFill("solid", fgColor="1F3864")
HDR_FONT  = Font(color="FFFFFF", bold=True, name="Calibri", size=10)
ALT_FILL  = PatternFill("solid", fgColor="DCE6F1")
NORM_FONT = Font(name="Calibri", size=10)
TITLE_FONT= Font(bold=True, name="Calibri", size=12, color="1F3864")
ACCENT    = PatternFill("solid", fgColor="2E75B6")
ACCENT_F  = Font(color="FFFFFF", bold=True, name="Calibri", size=10)
GREEN_F   = PatternFill("solid", fgColor="E2EFDA")
RED_F     = PatternFill("solid", fgColor="FCE4D6")

thin = Side(style="thin", color="BFBFBF")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)

def style_header_row(ws, row, ncols):
    for col in range(1, ncols + 1):
        c = ws.cell(row=row, column=col)
        c.fill = HDR_FILL; c.font = HDR_FONT
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = BORDER

def style_data_row(ws, row, ncols, alt=False):
    fill = ALT_FILL if alt else PatternFill()
    for col in range(1, ncols + 1):
        c = ws.cell(row=row, column=col)
        if alt: c.fill = fill
        c.font = NORM_FONT
        c.alignment = Alignment(horizontal="right")
        c.border = BORDER

def set_col_widths(ws, widths):
    for col, w in enumerate(widths, 1):
        ws.column_dimensions[ws.cell(1, col).column_letter].width = w

def write_df_to_sheet(ws, df, title, num_fmt="#,##0.0000", freeze=True,
                      index_label="Date"):
    ws.append([title])
    ws["A1"].font = TITLE_FONT
    ws.append([])
    hdr_row = 3
    ws.append([index_label] + list(df.columns))
    style_header_row(ws, hdr_row, len(df.columns) + 1)
    for i, (idx, row) in enumerate(df.iterrows()):
        r = hdr_row + 1 + i
        ws.append([str(idx.date())] + list(row.values))
        style_data_row(ws, r, len(df.columns) + 1, alt=(i % 2 == 1))
        for col in range(2, len(df.columns) + 2):
            ws.cell(r, col).number_format = num_fmt
    ws.column_dimensions["A"].width = 14
    for col in range(2, len(df.columns) + 2):
        ws.column_dimensions[ws.cell(hdr_row, col).column_letter].width = 12
    if freeze:
        ws.freeze_panes = ws.cell(hdr_row + 1, 2)


def build_excel(yields: pd.DataFrame, changes: pd.DataFrame,
                Sigma: np.ndarray, vals: np.ndarray, vecs: np.ndarray,
                pcs: pd.DataFrame, dv01s: dict, dv01_vec: np.ndarray,
                weights: np.ndarray, var_result: dict):

    wb = Workbook()
    wb.remove(wb.active)   # remove default sheet

    # ── Sheet 1: Raw Yields ────────────────────────────────────────────────
    ws1 = wb.create_sheet("1_Raw_Yields")
    write_df_to_sheet(ws1, yields, "US Treasury CMT Yields (%) — Daily", num_fmt="0.0000")

    # ── Sheet 2: Daily Changes (bps) ──────────────────────────────────────
    ws2 = wb.create_sheet("2_Daily_Changes_bps")
    write_df_to_sheet(ws2, changes, "Daily Yield Changes (basis points)", num_fmt="0.00")

    # ── Sheet 3: Covariance Matrix ─────────────────────────────────────────
    ws3 = wb.create_sheet("3_Covariance_Matrix")
    ws3.append(["Covariance Matrix of Daily Yield Changes (bps²)"])
    ws3["A1"].font = TITLE_FONT
    ws3.append([])
    ws3.append([""] + TENORS)
    style_header_row(ws3, 3, len(TENORS) + 1)
    for i, tenor in enumerate(TENORS):
        row = [tenor] + [round(Sigma[i, j], 6) for j in range(len(TENORS))]
        ws3.append(row)
        r = 4 + i
        ws3.cell(r, 1).fill = HDR_FILL; ws3.cell(r, 1).font = HDR_FONT
        ws3.cell(r, 1).border = BORDER
        for col in range(2, len(TENORS) + 2):
            c = ws3.cell(r, col)
            c.number_format = "0.000000"
            c.border = BORDER
            c.font = NORM_FONT
            c.alignment = Alignment(horizontal="right")
            if i % 2: c.fill = ALT_FILL
    for col in range(1, len(TENORS) + 2):
        ws3.column_dimensions[ws3.cell(1, col).column_letter].width = 13

    # ── Sheet 4: Eigenvalues & Explained Variance ─────────────────────────
    ws4 = wb.create_sheet("4_Eigenvalues")
    ws4.append(["Eigenvalues & Explained Variance"])
    ws4["A1"].font = TITLE_FONT
    ws4.append([])
    headers = ["PC", "Eigenvalue", "Std Dev (bps)", "Var Explained (%)",
               "Cum Var Explained (%)"]
    ws4.append(headers)
    style_header_row(ws4, 3, len(headers))
    total_var   = vals.sum()
    cum_var     = 0
    pc_labels   = ["PC1 Level", "PC2 Slope", "PC3 Curvature"] + \
                  [f"PC{i+1}" for i in range(3, len(vals))]
    for i, (val, lbl) in enumerate(zip(vals, pc_labels)):
        pct     = val / total_var * 100
        cum_var += pct
        r = 4 + i
        ws4.append([lbl, round(val, 6), round(np.sqrt(val), 4),
                    round(pct, 4), round(cum_var, 4)])
        style_data_row(ws4, r, len(headers), alt=(i % 2))
        ws4.cell(r, 1).alignment = Alignment(horizontal="left")
        if i < 3:
            for col in range(1, len(headers) + 1):
                ws4.cell(r, col).fill = GREEN_F
    set_col_widths(ws4, [18, 14, 14, 18, 22])

    # ── Sheet 5: Eigenvectors (Factor Loadings) ────────────────────────────
    ws5 = wb.create_sheet("5_Eigenvectors")
    ws5.append(["Eigenvectors (Factor Loadings) — first 3 PCs"])
    ws5["A1"].font = TITLE_FONT
    ws5.append([])
    ws5.append(["Tenor"] + ["PC1 Level", "PC2 Slope", "PC3 Curvature"])
    style_header_row(ws5, 3, 4)
    for i, tenor in enumerate(TENORS):
        r = 4 + i
        ws5.append([tenor] + [round(vecs[i, k], 6) for k in range(3)])
        style_data_row(ws5, r, 4, alt=(i % 2))
        ws5.cell(r, 1).alignment = Alignment(horizontal="left")
    set_col_widths(ws5, [10, 15, 15, 18])

    # ── Sheet 6: PC Summary Statistics (2007–2026) ────────────────────────
    # Row layout: hdr=row3, PC1=row4, PC2=row5, PC3=row6
    # Col E = 1-day Std Dev  ← referenced by Sheet 8 as '6_PC_Summary'!E4/5/6
    ws6 = wb.create_sheet("6_PC_Summary")
    ws6.cell(1, 1, "Principal Component Summary Statistics  (2007 to 2026)").font = TITLE_FONT
    ws6.merge_cells("A1:K1")
    ws6.append([])
    pc_sum_hdrs = [
        "PC", "Eigenvalue (bps^2)", "% Var Explained", "Cum % Var",
        "1-Day Std Dev (bps)", "10-Day Std Dev (bps)",
        "Historical Mean (bps)", "Historical Min (bps)", "Historical Max (bps)",
        "Skewness", "Excess Kurtosis",
    ]
    ws6.append(pc_sum_hdrs)
    style_header_row(ws6, 3, len(pc_sum_hdrs))
    total_var_s6 = vals.sum()
    cum_s6 = 0.0
    pc_names = ["PC1 Level", "PC2 Slope", "PC3 Curvature"]
    for k in range(3):
        col_data = pcs.iloc[:, k].dropna()
        pct_s6   = vals[k] / total_var_s6 * 100
        cum_s6  += pct_s6
        std_1d   = float(col_data.std())
        r = 4 + k
        ws6.append([
            pc_names[k],
            round(vals[k], 6),
            round(pct_s6, 4),
            round(cum_s6, 4),
            round(std_1d, 6),
            round(std_1d * np.sqrt(10), 6),
            round(float(col_data.mean()), 6),
            round(float(col_data.min()), 4),
            round(float(col_data.max()), 4),
            round(float(stats.skew(col_data)), 4),
            round(float(stats.kurtosis(col_data)), 4),
        ])
        style_data_row(ws6, r, len(pc_sum_hdrs), alt=False)
        for col in range(1, len(pc_sum_hdrs) + 1):
            ws6.cell(r, col).fill = GREEN_F
            ws6.cell(r, col).font = NORM_FONT
        ws6.cell(r, 1).fill = HDR_FILL
        ws6.cell(r, 1).font = HDR_FONT
        ws6.cell(r, 1).alignment = Alignment(horizontal="left")
    pc_sum_widths = [18, 20, 18, 12, 20, 22, 22, 20, 20, 12, 16]
    col_letters_s6 = ["A","B","C","D","E","F","G","H","I","J","K"]
    for letter, w in zip(col_letters_s6, pc_sum_widths):
        ws6.column_dimensions[letter].width = w
    ws6.freeze_panes = "B4"

    # ── Sheet 7: Portfolio DV01 & Factor Risk Weights (formula-driven) ────
    # Cell map (cross-referenced by Sheet 8):
    #   Sec A inputs  : A4:F6  (hdr row4 | 2Y row5 | 10Y row6)
    #                   col A=Bond  B=Notional  C=Coupon(dec)  D=YTM(dec)
    #                   col E=Tenor(yrs)  F=Settlement(=TODAY())
    #   Sec B DV01    : A9:E11 (hdr row9 | 2Y row10 | 10Y row11)
    #                   col A=Bond  B=Maturity  C=PRICE  D=MDURATION  E=DV01
    #   Sec C weights : A14:F17 (hdr row14 | PC1 row15 | PC2 row16 | PC3 row17)
    #                   col A=PC  B=Load@2Y  C=Load@10Y  D=DV01@2Y  E=DV01@10Y  F=Risk Weight
    ws7 = wb.create_sheet("7_DV01_Risk_Weights")
    INPUT_FILL = PatternFill("solid", fgColor="FFF2CC")

    def sec_hdr(ws, row, text, ncols=6):
        ws.cell(row, 1, text).font = Font(bold=True, name="Calibri", size=11, color="1F3864")
        ws.cell(row, 1).fill = PatternFill("solid", fgColor="D6E4F0")
        if ncols > 1:
            ws.merge_cells(f"A{row}:{chr(64+ncols)}{row}")

    ws7.cell(1, 1, "Portfolio DV01 & Factor Risk Weights").font = TITLE_FONT
    ws7.merge_cells("A1:F1")
    ws7.append([])

    # ── Sec A: Bond Parameters ─────────────────────────────────────────────
    sec_hdr(ws7, 3, "SECTION A  |  Bond Parameters  (edit highlighted cells)")
    inp_hdrs = ["Bond", "Notional ($)", "Annual Coupon", "YTM / Last CMT", "Tenor (Yrs)", "Settlement"]
    for c, h in enumerate(inp_hdrs, 1):
        ws7.cell(4, c, h)
    style_header_row(ws7, 4, 6)

    ytm_2y  = float(yields.iloc[-1]["2Y"])  / 100   # as decimal
    ytm_10y = float(yields.iloc[-1]["10Y"]) / 100

    for row_num, (tenor_key, ytm_dec, mat_yrs) in enumerate(
            [("2Y", ytm_2y, 2), ("10Y", ytm_10y, 10)], start=5):
        face   = PORTFOLIO[tenor_key]
        coupon = COUPON_RATES[tenor_key]
        ws7.cell(row_num, 1, f"{tenor_key} T-Note")
        ws7.cell(row_num, 2, face).number_format = "#,##0"
        ws7.cell(row_num, 3, coupon).number_format = "0.00%"       # stored as decimal
        ws7.cell(row_num, 4, ytm_dec).number_format = "0.0000%"    # stored as decimal
        ws7.cell(row_num, 5, mat_yrs).number_format = "0"
        ws7.cell(row_num, 6, "=TODAY()").number_format = "YYYY-MM-DD"
        for c in range(1, 7):
            ws7.cell(row_num, c).font  = NORM_FONT
            ws7.cell(row_num, c).border = BORDER
            ws7.cell(row_num, c).alignment = Alignment(horizontal="right")
        ws7.cell(row_num, 1).alignment = Alignment(horizontal="left")
        for c in [2, 3, 4, 5]:
            ws7.cell(row_num, c).fill = INPUT_FILL

    # ── Sec B: DV01 (Excel PRICE & MDURATION) ─────────────────────────────
    ws7.append([])
    sec_hdr(ws7, 8, "SECTION B  |  DV01 Calculation  (Excel PRICE & MDURATION functions)")
    dv01_hdrs = ["Bond", "Maturity Date", "Full Price (per $100 Face)",
                 "Modified Duration (yrs)", "DV01  (USD per bp)"]
    for c, h in enumerate(dv01_hdrs, 1):
        ws7.cell(9, c, h)
    style_header_row(ws7, 9, 5)

    #  DV01 = (Notional/100) * PRICE(settlement,maturity,coupon,ytm,100,2)
    #                         * MDURATION(settlement,maturity,coupon,ytm,2) / 10,000
    dv01_formula_rows = [
        # (row, bond label, input_row for Sec A)
        (10, "2Y T-Note",  5),
        (11, "10Y T-Note", 6),
    ]
    for r, label, inp_r in dv01_formula_rows:
        ws7.cell(r, 1, label).font  = NORM_FONT
        ws7.cell(r, 1).border = BORDER
        ws7.cell(r, 1).alignment = Alignment(horizontal="left")
        ws7.cell(r, 2, f"=EDATE($F${inp_r},$E${inp_r}*12)")
        ws7.cell(r, 2).number_format = "YYYY-MM-DD"
        ws7.cell(r, 3, f"=PRICE($F${inp_r},B{r},$C${inp_r},$D${inp_r},100,2)")
        ws7.cell(r, 3).number_format = "0.0000"
        ws7.cell(r, 4, f"=MDURATION($F${inp_r},B{r},$C${inp_r},$D${inp_r},2)")
        ws7.cell(r, 4).number_format = "0.0000"
        ws7.cell(r, 5, f"=$B${inp_r}/100*C{r}*D{r}/10000")
        ws7.cell(r, 5).number_format = "#,##0.00"
        for c in range(2, 6):
            ws7.cell(r, c).font   = NORM_FONT
            ws7.cell(r, c).border = BORDER
            ws7.cell(r, c).alignment = Alignment(horizontal="right")

    # ── Sec C: Factor Risk Weights ────────────────────────────────────────
    ws7.append([])
    sec_hdr(ws7, 13,
            "SECTION C  |  Factor Risk Weights  =  DV01_2Y x Loading_2Y  +  DV01_10Y x Loading_10Y")
    rw_hdrs = ["PC Factor", "PC Loading @ 2Y (*)",
               "PC Loading @ 10Y (*)", "DV01 @ 2Y ($)", "DV01 @ 10Y ($)",
               "Risk Weight  ($ / bp of PC)"]
    for c, h in enumerate(rw_hdrs, 1):
        ws7.cell(14, c, h)
    style_header_row(ws7, 14, 6)

    # Sheet 5 eigenvector layout: hdr row3 data from row4
    # 2Y  = TENORS index 3 → row 4+3 = 7  ; cols B=PC1, C=PC2, D=PC3
    # 10Y = TENORS index 7 → row 4+7 = 11
    ev_col_letters = ["B", "C", "D"]   # PC1, PC2, PC3
    for k, nm in enumerate(pc_names):
        r   = 15 + k
        evc = ev_col_letters[k]
        ws7.cell(r, 1, nm).font  = NORM_FONT
        ws7.cell(r, 1).border = BORDER
        ws7.cell(r, 1).alignment = Alignment(horizontal="left")
        ws7.cell(r, 2, f"='5_Eigenvectors'!{evc}7").number_format  = "0.000000"
        ws7.cell(r, 3, f"='5_Eigenvectors'!{evc}11").number_format = "0.000000"
        ws7.cell(r, 4, "=E10").number_format = "#,##0.00"
        ws7.cell(r, 5, "=E11").number_format = "#,##0.00"
        ws7.cell(r, 6, f"=B{r}*D{r}+C{r}*E{r}").number_format = "#,##0.00"
        for c in range(2, 7):
            ws7.cell(r, c).font   = NORM_FONT
            ws7.cell(r, c).border = BORDER
            ws7.cell(r, c).alignment = Alignment(horizontal="right")
        if k % 2:
            for c in range(1, 7):
                ws7.cell(r, c).fill = ALT_FILL
        ws7.cell(r, 6).font = Font(bold=True, name="Calibri", size=10)

    note_r = 18
    ws7.cell(note_r, 1,
             "(*) PC eigenvector loadings from Sheet 5_Eigenvectors  |  "
             "DV01 = (Notional/100) x PRICE(settlement,maturity,coupon,ytm,100,2) x MDURATION(...) / 10,000"
             ).font = Font(italic=True, name="Calibri", size=9, color="595959")
    ws7.merge_cells(f"A{note_r}:F{note_r}")

    for letter, w in zip(["A","B","C","D","E","F"], [18, 24, 28, 26, 14, 28]):
        ws7.column_dimensions[letter].width = w
    ws7.freeze_panes = "B5"

    # ── Sheet 8: VaR Results (formula-driven) ─────────────────────────────
    # Cross-sheet refs used:
    #   '6_PC_Summary'!E4/5/6  = 1-day std dev for PC1/2/3
    #   '7_DV01_Risk_Weights'!F15/16/17 = risk weights for PC1/2/3
    # Local anchor cells:
    #   B4 = confidence (0.99)  |  B5 = horizon (10)
    #   B9/B10/B11 = 1-day std dev (from Sheet6)
    #   D9/D10/D11 = risk weights (from Sheet7)
    #   E9/E10/E11 = 10-day $ risk std dev
    ws8 = wb.create_sheet("8_VaR_Results")

    def sec_hdr8(ws, row, text, ncols=5):
        ws.cell(row, 1, text).font = Font(bold=True, name="Calibri", size=11, color="1F3864")
        ws.cell(row, 1).fill = PatternFill("solid", fgColor="D6E4F0")
        ws.merge_cells(f"A{row}:{chr(64+ncols)}{row}")

    ws8.cell(1, 1, "Value-at-Risk (VaR)  —  PCA Factor Model").font = TITLE_FONT
    ws8.merge_cells("A1:E1")
    ws8.append([])

    # Sec A: Parameters
    sec_hdr8(ws8, 3, "SECTION A  |  VaR Parameters  (edit highlighted cells)", 3)
    for r, (label, val, fmt) in enumerate(
            [("Confidence Level", 0.99, "0%"), ("Horizon (days)", 10, "0")], start=4):
        ws8.cell(r, 1, label).font  = Font(bold=True, name="Calibri", size=10)
        ws8.cell(r, 1).border = BORDER
        ws8.cell(r, 2, val).number_format = fmt
        ws8.cell(r, 2).fill   = INPUT_FILL
        ws8.cell(r, 2).border = BORDER
        ws8.cell(r, 2).font   = NORM_FONT
        ws8.cell(r, 2).alignment = Alignment(horizontal="right")

    # Sec B: PC Factor Statistics (formula references to Sheets 6 & 7)
    sec_hdr8(ws8, 7, "SECTION B  |  PC Factor Statistics  (linked from Sheets 6 & 7)")
    pc_stat_hdrs = ["PC Factor", "1-Day Std Dev (bps)", "10-Day Std Dev (bps)",
                    "Risk Weight ($ / bp)", "10-Day $ Risk Std Dev"]
    for c, h in enumerate(pc_stat_hdrs, 1):
        ws8.cell(8, c, h)
    style_header_row(ws8, 8, 5)

    s6_std_rows = [4, 5, 6]     # rows in Sheet 6 for PC1/2/3 1-day std (col E)
    s7_rw_rows  = [15, 16, 17]  # rows in Sheet 7 for risk weights (col F)
    for k, nm in enumerate(pc_names):
        r = 9 + k
        ws8.cell(r, 1, nm).font  = NORM_FONT
        ws8.cell(r, 1).border = BORDER
        ws8.cell(r, 1).alignment = Alignment(horizontal="left")
        ws8.cell(r, 2, f"='6_PC_Summary'!E{s6_std_rows[k]}").number_format = "0.0000"
        ws8.cell(r, 3, f"=B{r}*SQRT($B$5)").number_format = "0.0000"
        ws8.cell(r, 4, f"='7_DV01_Risk_Weights'!F{s7_rw_rows[k]}").number_format = "#,##0.00"
        ws8.cell(r, 5, f"=ABS(C{r}*D{r})").number_format = "#,##0.00"
        for c in range(2, 6):
            ws8.cell(r, c).font   = NORM_FONT
            ws8.cell(r, c).border = BORDER
            ws8.cell(r, c).alignment = Alignment(horizontal="right")
        if k % 2:
            for c in range(1, 6):
                ws8.cell(r, c).fill = ALT_FILL

    # Sec C: VaR Computation — all three methods
    sec_hdr8(ws8, 13,
             "SECTION C  |  VaR Computation  —  Three Methods Compared")

    # Intermediate building blocks (rows 14-16)
    intermediate = [
        ("Portfolio Variance  (USD²)   =  Σ (Risk Weight × 10-Day Std Dev)²",
         "=E9^2+E10^2+E11^2",                         "#,##0.00",   False),
        ("Portfolio Std Dev  (USD)   =  √(Variance)",
         "=SQRT(B14)",                                 "#,##0.00",   False),
        ("z-score  at Confidence  =  NORM.S.INV(Confidence)",
         "=NORM.S.INV($B$4)",                          "0.0000",     False),
    ]
    for i, (label, val, fmt, _) in enumerate(intermediate):
        r = 14 + i
        ws8.cell(r, 1, label).font = NORM_FONT
        ws8.cell(r, 2, val).number_format = fmt
        ws8.cell(r, 1).border = BORDER
        ws8.cell(r, 2).border = BORDER
        ws8.cell(r, 2).font   = NORM_FONT
        ws8.cell(r, 2).alignment = Alignment(horizontal="right")
        if i % 2:
            for c in [1, 2]: ws8.cell(r, c).fill = ALT_FILL

    # Blank separator
    ws8.cell(17, 1, "").border = BORDER
    ws8.cell(17, 2, "").border = BORDER

    # Row 18: Method 1 — Analytical
    ws8.cell(18, 1,
             "METHOD 1  |  Analytical VaR  (Closed-Form, Normal Assumption)")
    ws8.cell(18, 2,
             "=NORM.S.INV($B$4)*SQRT(($D$9*$C$9)^2+($D$10*$C$10)^2+($D$11*$C$11)^2)"
             ).number_format = '"$"#,##0'
    ws8.cell(18, 3,
             "Formula: z × √[Σ(wₖ × σₖ × √T)²]   —   "
             "Exact result under normality. Monte Carlo must converge to this."
             ).font = Font(italic=True, name="Calibri", size=9, color="595959")
    ws8.merge_cells("C18:E18")
    for c in [1, 2]:
        ws8.cell(18, c).fill   = ALT_FILL
        ws8.cell(18, c).border = BORDER
        ws8.cell(18, c).font   = Font(bold=True, name="Calibri", size=10)
    ws8.cell(18, 2).alignment = Alignment(horizontal="right")
    ws8.cell(18, 3).border = BORDER

    # Row 19: Method 2 — Parametric Monte Carlo
    ws8.cell(19, 1,
             "METHOD 2  |  Parametric Monte Carlo VaR  (Normal Draws, 10,000 Paths)")
    ws8.cell(19, 2, round(var_result["var_mc_usd"])).number_format = '"$"#,##0'
    ws8.cell(19, 3,
             f"Draws PC moves from N(0, σₖ×√T). Same normal assumption as Method 1 "
             f"— agrees to within sampling error (~{abs(var_result['var_mc_usd']-var_result['var_analytical_usd'])/var_result['var_analytical_usd']*100:.1f}%). "
             f"Paths computed in Python (see Sheet 10 for full P&L audit trail)."
             ).font = Font(italic=True, name="Calibri", size=9, color="595959")
    ws8.merge_cells("C19:E19")
    for c in [1, 2]:
        ws8.cell(19, c).fill   = ALT_FILL
        ws8.cell(19, c).border = BORDER
        ws8.cell(19, c).font   = Font(bold=True, name="Calibri", size=10)
    ws8.cell(19, 2).alignment = Alignment(horizontal="right")
    ws8.cell(19, 3).border = BORDER

    # Row 20: Method 3 — True Historical Simulation (the only one that matters)
    hist_excess = (var_result["var_hist_usd"] / var_result["var_analytical_usd"] - 1) * 100
    ws8.cell(20, 1,
             "METHOD 3  |  Historical Simulation VaR  (Empirical Distribution, No Normality Assumption)")
    ws8.cell(20, 2, round(var_result["var_hist_usd"])).number_format = '"$"#,##0'
    ws8.cell(20, 3,
             f"Uses actual {len(var_result['pnl_hist_10d']):,} daily PC observations × √T. "
             f"Empirical P&L excess kurtosis = {var_result['pnl_kurt']:.2f} (0=normal). "
             f"Fat tails inflate VaR by {hist_excess:+.1f}% vs normal assumption. "
             f"See Sheet 10_HistSim_PnL for full date-by-date P&L."
             ).font = Font(italic=True, name="Calibri", size=9, color="595959")
    ws8.merge_cells("C20:E20")
    for c in [1, 2]:
        ws8.cell(20, c).fill   = RED_F
        ws8.cell(20, c).border = BORDER
        ws8.cell(20, c).font   = Font(bold=True, name="Calibri", size=11, color="C00000")
    ws8.cell(20, 2).alignment = Alignment(horizontal="right")
    ws8.cell(20, 3).border = BORDER

    # Row 21: empirical distribution stats
    ws8.cell(21, 1,
             f"  Daily P&L Excess Kurtosis = {var_result['pnl_kurt']:.2f}   |   "
             f"Skewness = {var_result['pnl_skew']:.2f}   |   "
             f"Normal assumption kurtosis = 0.00 by definition"
             ).font = Font(italic=True, name="Calibri", size=9, color="7F7F7F")
    ws8.cell(21, 1).border = BORDER
    ws8.merge_cells("A21:E21")

    # Sec D: Standalone PC VaR Contributions (Historical Simulation)
    sec_hdr8(ws8, 23, "SECTION D  |  Standalone PC VaR Contributions  (Historical Simulation, Undiversified)")
    contrib_hdrs = ["PC Factor", "Standalone VaR — Hist Sim (USD)",
                    "% of Sum (undiversified)", "Standalone VaR — Analytical (USD)"]
    for c, h in enumerate(contrib_hdrs, 1):
        ws8.cell(24, c, h)
    style_header_row(ws8, 24, 4)

    for k, nm in enumerate(pc_names):
        r = 25 + k
        ws8.cell(r, 1, nm).font  = NORM_FONT
        ws8.cell(r, 1).border = BORDER
        ws8.cell(r, 1).alignment = Alignment(horizontal="left")
        ws8.cell(r, 2, round(var_result["contrib_hist"][k])).number_format = '"$"#,##0'
        ws8.cell(r, 3, f"=B{r}/SUM($B$25:$B$27)").number_format = "0.0%"
        ws8.cell(r, 4, f"=NORM.S.INV($B$4)*E{9+k}").number_format = '"$"#,##0'
        for c in [2, 3, 4]:
            ws8.cell(r, c).font   = NORM_FONT
            ws8.cell(r, c).border = BORDER
            ws8.cell(r, c).alignment = Alignment(horizontal="right")
        if k % 2:
            for c in range(1, 5): ws8.cell(r, c).fill = ALT_FILL

    # Diversified total row
    r_tot = 28
    ws8.cell(r_tot, 1, "Total (Diversified)  — Historical Simulation").font = Font(
        bold=True, name="Calibri", size=10)
    ws8.cell(r_tot, 2, round(var_result["var_hist_usd"])).number_format = '"$"#,##0'
    ws8.cell(r_tot, 3, f"=B{r_tot}/SUM($B$25:$B$27)").number_format = "0.0%"
    ws8.cell(r_tot, 4,
             "=NORM.S.INV($B$4)*SQRT(E9^2+E10^2+E11^2)").number_format = '"$"#,##0'
    for c in range(1, 5):
        ws8.cell(r_tot, c).border = BORDER
        ws8.cell(r_tot, c).fill   = GREEN_F
    ws8.cell(r_tot, 2).font = Font(bold=True, name="Calibri", size=10)
    ws8.cell(r_tot, 2).alignment = Alignment(horizontal="right")
    ws8.cell(r_tot, 3).font = Font(bold=True, name="Calibri", size=10)
    ws8.cell(r_tot, 3).alignment = Alignment(horizontal="right")
    ws8.cell(r_tot, 4).font = Font(bold=True, name="Calibri", size=10)
    ws8.cell(r_tot, 4).alignment = Alignment(horizontal="right")

    ws8.cell(29, 1,
             "Note: Diversification benefit = Sum(standalone) - Diversified VaR. "
             "PCs are orthogonal so analytical variance is additive. "
             "Historical sim diversification arises from actual observed correlations being near-zero."
             ).font = Font(italic=True, name="Calibri", size=9, color="595959")
    ws8.merge_cells("A29:E29")

    for letter, w in zip(["A","B","C","D","E"], [42, 24, 14, 42, 18]):
        ws8.column_dimensions[letter].width = w

    # ── Sheet 9: PC Correlation Check ─────────────────────────────────────
    ws9 = wb.create_sheet("9_PC_Correlations")
    ws9.append(["PC Correlation Matrix (should be identity — orthogonality check)"])
    ws9["A1"].font = TITLE_FONT
    ws9.append([])
    pc_corr = np.corrcoef(var_result["sim_moves"].T)
    pc_names_short = ["PC1", "PC2", "PC3"]
    ws9.append([""] + pc_names_short)
    style_header_row(ws9, 3, 4)
    for i, nm in enumerate(pc_names_short):
        r = 4 + i
        ws9.append([nm] + [round(pc_corr[i, j], 6) for j in range(3)])
        style_data_row(ws9, r, 4, alt=(i % 2))
        ws9.cell(r, 1).fill = HDR_FILL; ws9.cell(r, 1).font = HDR_FONT
    set_col_widths(ws9, [10, 14, 14, 14])

    # ── Sheet 10: Historical Simulation P&L — Full Audit Trail ─────────────
    # Every trading day's actual PC moves → P&L → scaled to 10-day horizon.
    # Sorted worst-to-best. Tail rows (below VaR threshold) highlighted red.
    ws10 = wb.create_sheet("10_HistSim_PnL")
    ws10.cell(1, 1,
              f"Historical Simulation P&L — Audit Trail   "
              f"(sorted worst-to-best  |  {HORIZON}-day scaled via sqrt(T)  |  "
              f"{CONFIDENCE*100:.0f}% VaR tail highlighted in red)"
              ).font = TITLE_FONT
    ws10.merge_cells("A1:H1")
    ws10.append([])

    pnl_dates = var_result["pnl_hist_dates"]
    pnl_10d   = var_result["pnl_hist_10d"]
    daily_pnl = var_result["daily_pnl"]

    hs_df = pd.DataFrame({
        "Date":                       pnl_dates,
        "PC1 Move (bps)":             pcs.values[:, 0],
        "PC2 Move (bps)":             pcs.values[:, 1],
        "PC3 Move (bps)":             pcs.values[:, 2],
        "Daily P&L ($)":              daily_pnl,
        f"{HORIZON}-Day P&L ($)":     pnl_10d,
    }).sort_values(f"{HORIZON}-Day P&L ($)", ascending=True).reset_index(drop=True)

    n_obs    = len(hs_df)
    var_cut  = np.percentile(pnl_10d, (1 - CONFIDENCE) * 100)
    n_breach = int(np.sum(pnl_10d <= var_cut))

    hs_hdrs = ["Rank (Worst=1)", "Date", "PC1 Move (bps)", "PC2 Move (bps)",
               "PC3 Move (bps)", "Daily P&L ($)", f"{HORIZON}-Day Scaled P&L ($)",
               "Cum. Percentile (%)"]
    for c, h in enumerate(hs_hdrs, 1):
        ws10.cell(3, c, h)
    style_header_row(ws10, 3, len(hs_hdrs))

    # Banner explaining the tail
    sep_r = 4
    ws10.cell(sep_r, 1,
              f"  <<< VaR TAIL  |  {n_breach} worst days out of {n_obs:,} observations  "
              f"({n_breach/n_obs*100:.2f}% = 1 - {CONFIDENCE*100:.0f}% confidence)  "
              f"|  VaR cutoff = ${abs(var_cut):,.0f}  >>>"
              ).font = Font(bold=True, name="Calibri", size=10, color="C00000")
    ws10.cell(sep_r, 1).fill = PatternFill("solid", fgColor="FCE4D6")
    ws10.merge_cells(f"A{sep_r}:H{sep_r}")

    for i, (_, row_data) in enumerate(hs_df.iterrows()):
        r      = sep_r + 1 + i
        rank   = i + 1
        spnl   = round(float(row_data[f"{HORIZON}-Day P&L ($)"]), 2)
        dpnl   = round(float(row_data["Daily P&L ($)"]), 2)
        in_tail = spnl <= var_cut

        ws10.cell(r, 1, rank)
        ws10.cell(r, 2, str(row_data["Date"].date()))
        ws10.cell(r, 3, round(float(row_data["PC1 Move (bps)"]), 3)).number_format = "0.000"
        ws10.cell(r, 4, round(float(row_data["PC2 Move (bps)"]), 3)).number_format = "0.000"
        ws10.cell(r, 5, round(float(row_data["PC3 Move (bps)"]), 3)).number_format = "0.000"
        ws10.cell(r, 6, dpnl).number_format = '"$"#,##0'
        ws10.cell(r, 7, spnl).number_format = '"$"#,##0'
        ws10.cell(r, 8, round(rank / n_obs * 100, 4)).number_format = "0.0000%"

        fill = RED_F if in_tail else (ALT_FILL if i % 2 else PatternFill())
        for c in range(1, 9):
            ws10.cell(r, c).font   = NORM_FONT
            ws10.cell(r, c).border = BORDER
            ws10.cell(r, c).alignment = Alignment(horizontal="right")
            if in_tail or (i % 2 and not in_tail):
                ws10.cell(r, c).fill = fill
        ws10.cell(r, 2).alignment = Alignment(horizontal="left")
        # Bold the exact VaR boundary row
        if rank == n_breach:
            for c in range(1, 9):
                ws10.cell(r, c).font = Font(bold=True, name="Calibri", size=10,
                                            color="C00000")

    for letter, w in zip(["A","B","C","D","E","F","G","H"],
                         [16, 14, 18, 18, 18, 16, 22, 20]):
        ws10.column_dimensions[letter].width = w
    ws10.freeze_panes = "A5"

    wb.save(OUTPUT_FILE)
    print(f"\n  Saved: {OUTPUT_FILE}")


# ──────────────────────────────────────────────────────────────────────────────
# PLOTS (saved as PNG, then embedded — openpyxl image insertion requires pillow)
# ──────────────────────────────────────────────────────────────────────────────
def make_plots(yields: pd.DataFrame, changes: pd.DataFrame, vals: np.ndarray,
               vecs: np.ndarray, pcs: pd.DataFrame, pnl: np.ndarray,
               var_usd: float):

    fig_dir = "pca_plots"
    os.makedirs(fig_dir, exist_ok=True)

    # 1. Yield curves over time (sample every ~252 days)
    fig, ax = plt.subplots(figsize=(12, 5))
    sample  = yields.iloc[::252]
    cm      = plt.cm.viridis(np.linspace(0, 1, len(sample)))
    for i, (dt, row) in enumerate(sample.iterrows()):
        ax.plot(TENOR_YR, row.values, color=cm[i], linewidth=1.2,
                label=str(dt.date()))
    ax.set_title("US Treasury Yield Curves (sampled annually)", fontsize=13, fontweight="bold")
    ax.set_xlabel("Maturity (years)"); ax.set_ylabel("Yield (%)")
    ax.legend(fontsize=7, ncol=3, loc="upper left")
    ax.grid(True, alpha=0.3)
    plt.tight_layout()
    fig.savefig(f"{fig_dir}/01_yield_curves.png", dpi=150)
    plt.close(fig)

    # 2. Explained variance scree plot
    total_var = vals.sum()
    pct = vals / total_var * 100
    cum = np.cumsum(pct)
    fig, ax = plt.subplots(figsize=(9, 4))
    ax.bar(range(1, len(vals) + 1), pct, color="#2E75B6", edgecolor="white")
    ax2 = ax.twinx()
    ax2.plot(range(1, len(vals) + 1), cum, "o-", color="#C00000", linewidth=2)
    ax2.axhline(95, color="gray", linestyle="--", linewidth=1)
    ax.set_title("Scree Plot — Explained Variance", fontsize=13, fontweight="bold")
    ax.set_xlabel("Principal Component"); ax.set_ylabel("Variance Explained (%)")
    ax2.set_ylabel("Cumulative Variance Explained (%)", color="#C00000")
    ax.grid(True, alpha=0.3)
    plt.tight_layout()
    fig.savefig(f"{fig_dir}/02_scree_plot.png", dpi=150)
    plt.close(fig)

    # 3. Factor loadings (eigenvectors)
    fig, ax = plt.subplots(figsize=(9, 5))
    colors = ["#2E75B6", "#C00000", "#70AD47"]
    labels = ["PC1 Level", "PC2 Slope", "PC3 Curvature"]
    for k in range(3):
        ax.plot(TENOR_YR, vecs[:, k], "o-", color=colors[k], linewidth=2,
                markersize=6, label=labels[k])
    ax.axhline(0, color="black", linewidth=0.8)
    ax.set_title("Factor Loadings (Eigenvectors)", fontsize=13, fontweight="bold")
    ax.set_xlabel("Maturity (years)"); ax.set_ylabel("Loading")
    ax.legend(); ax.grid(True, alpha=0.3)
    plt.tight_layout()
    fig.savefig(f"{fig_dir}/03_factor_loadings.png", dpi=150)
    plt.close(fig)

    # 4. PC time series
    fig, axes = plt.subplots(3, 1, figsize=(12, 8), sharex=True)
    for k, (pc_col, color, lbl) in enumerate(
            zip(pcs.columns, colors, labels)):
        axes[k].fill_between(pcs.index, pcs[pc_col], alpha=0.4, color=color)
        axes[k].plot(pcs.index, pcs[pc_col], color=color, linewidth=0.6)
        axes[k].set_ylabel(f"{lbl}\n(bps)")
        axes[k].axhline(0, color="black", linewidth=0.6)
        axes[k].grid(True, alpha=0.3)
    axes[0].set_title("Daily PC Realizations", fontsize=13, fontweight="bold")
    plt.tight_layout()
    fig.savefig(f"{fig_dir}/04_pc_timeseries.png", dpi=150)
    plt.close(fig)

    # 5. P&L distribution: historical (empirical) vs normal overlay
    fig, ax = plt.subplots(figsize=(11, 5))
    # Historical empirical distribution
    ax.hist(pnl, bins=120, color="#2E75B6", edgecolor="white", alpha=0.75,
            density=True, label="Historical Simulation (empirical)")
    # Normal distribution overlay using same std dev
    pnl_std = float(np.std(pnl))
    x_range = np.linspace(pnl.min(), pnl.max(), 500)
    normal_pdf = stats.norm.pdf(x_range, 0, pnl_std)
    ax.plot(x_range, normal_pdf, color="#70AD47", linewidth=2.5, linestyle="-",
            label=f"Normal fit  (same std dev = ${pnl_std:,.0f})")
    ax.axvline(-var_usd, color="#C00000", linewidth=2.5, linestyle="--",
               label=f"Historical Sim VaR (99%) = ${var_usd:,.0f}")
    ax.set_title("10-Day P&L: Empirical Distribution vs Normal Fit",
                 fontsize=13, fontweight="bold")
    ax.set_xlabel("P&L (USD)"); ax.set_ylabel("Probability Density")
    ax.xaxis.set_major_formatter(mticker.FuncFormatter(lambda x, _: f"${x:,.0f}"))
    ax.legend(fontsize=9); ax.grid(True, alpha=0.3)
    plt.tight_layout()
    fig.savefig(f"{fig_dir}/05_pnl_distribution.png", dpi=150)
    plt.close(fig)

    print(f"  Plots saved to ./{fig_dir}/")
    return fig_dir


# ──────────────────────────────────────────────────────────────────────────────
# MAIN
# ──────────────────────────────────────────────────────────────────────────────
def main():
    print("=" * 60)
    print("  Treasury Yield PCA Factor Model & VaR")
    print("=" * 60)

    print("\n[1/7] Fetching FRED data …")
    raw = fetch_all()

    print("\n[2/7] Cleaning data (removing holidays & NaNs) …")
    yields  = clean(raw)
    print(f"  {len(yields)} business days  |  "
          f"{yields.index[0].date()} to {yields.index[-1].date()}")

    print("\n[3/7] Computing daily yield changes (bps) …")
    changes = yield_changes(yields)
    print(f"  Change matrix: {changes.shape}")

    print("\n[4/7] Eigendecomposition of covariance matrix …")
    vals, vecs, Sigma = eigen(changes)
    total_var = vals.sum()
    for i in range(3):
        print(f"  PC{i+1}: eigenvalue={vals[i]:.4f}  "
              f"({vals[i]/total_var*100:.2f}% variance explained)")

    print("\n[5/7] Extracting principal components …")
    pcs = compute_pcs(changes, vecs)

    print("\n[6/7] Computing DV01 & factor risk weights …")
    dv01s  = compute_dv01s(yields.iloc[-1])
    weights, dv01_vec = factor_risk_weights(dv01s, vecs)
    for tenor, dv01 in dv01s.items():
        print(f"  {tenor}: DV01 = ${dv01:,.2f} / bp")
    for k, (nm, w) in enumerate(zip(["Level","Slope","Curvature"], weights)):
        print(f"  PC{k+1} {nm} risk weight: ${w:,.2f}")

    print("\n[7/7] Computing VaR (99%, 10-day) — three methods …")
    var_result = compute_var(pcs, weights, CONFIDENCE, HORIZON, N_SIM)
    print(f"  Method 1 Analytical VaR        : ${var_result['var_analytical_usd']:>12,.0f}  (normal, closed-form)")
    print(f"  Method 2 Parametric MC VaR     : ${var_result['var_mc_usd']:>12,.0f}  (normal draws, 10k paths)")
    print(f"  Method 3 Historical Sim VaR    : ${var_result['var_hist_usd']:>12,.0f}  (empirical, fat tails)")
    print(f"  Daily P&L excess kurtosis      : {var_result['pnl_kurt']:>12.2f}  (0 = normal)")

    print("\n[+]  Generating plots …")
    make_plots(yields, changes, vals, vecs, pcs,
               var_result["pnl_hist_10d"], var_result["var_hist_usd"])

    print("\n[+]  Writing Excel workbook …")
    build_excel(yields, changes, Sigma, vals, vecs, pcs,
                dv01s, dv01_vec, weights, var_result)

    print("\n" + "=" * 60)
    print("  DONE")
    print("=" * 60)


if __name__ == "__main__":
    main()
